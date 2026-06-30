"""
phases/phase1/file_operations.py
File-system operations and report generation for the Phase 1 ingestor.

Responsibilities:
  • execute_copies()  — copy or move matched folders to the destination
  • generate_report() — write reconciliation_*.xlsx + missing_records_*.txt
"""
from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path
from typing import Callable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from phases.phase1.folder_classifier import MatchResult, FolderEntry
from phases.phase1.mappings import DEST_LABELS


# ── openpyxl style helpers ────────────────────────────────────────────────────

def _fill(hex_c: str)  -> PatternFill: return PatternFill("solid", fgColor=hex_c)
def _font(bold=False, colour="000000", size=10) -> Font:
    return Font(name="Arial", bold=bold, color=colour, size=size)
def _border() -> Border:
    s = Side(style="thin", color="D0D7DE")
    return Border(left=s, right=s, top=s, bottom=s)
def _hdr(ws, headers: list[str], row: int = 1) -> None:
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = _fill("1A2035"); cell.font = _font(bold=True, colour="ECEFF1")
        cell.border = _border()
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28
def _col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Copy / move engine ────────────────────────────────────────────────────────

def execute_copies(
    results:        list[MatchResult],
    dest_root:      Path,
    program_filter: str,
    do_move:        bool = False,
    on_item:        Optional[Callable[[MatchResult, bool, Optional[str]], None]] = None,
) -> tuple[int, int]:
    """
    Copies (or moves) every matched folder to dest_root/<label>/.

    Args:
        results:        Full match results list.
        dest_root:      Root destination directory.
        program_filter: Used to pick the sub-folder label.
        do_move:        True = shutil.move; False = shutil.copytree.
        on_item:        Optional callback(result, ok, error_msg) per item.

    Returns:
        (success_count, error_count)
    """
    matched = [r for r in results if r.folder is not None]
    if not matched:
        return 0, 0

    label    = DEST_LABELS.get(program_filter, "Graduates")
    dest_dir = dest_root / label
    dest_dir.mkdir(parents=True, exist_ok=True)

    ok = err = 0
    for r in matched:
        target     = dest_dir / r.folder.path.name
        r.dest_path = target

        if target.exists():
            if on_item:
                on_item(r, False, "destination already exists — skipped")
            ok += 1   # not an error, just pre-existing
            continue

        try:
            if do_move:
                shutil.move(str(r.folder.path), str(target))
            else:
                shutil.copytree(str(r.folder.path), str(target))
            ok += 1
            if on_item:
                on_item(r, True, None)
        except Exception as exc:
            err += 1
            if on_item:
                on_item(r, False, str(exc))

    return ok, err


# ── Report generation ─────────────────────────────────────────────────────────

def generate_report(
    results:           list[MatchResult],
    unmatched_folders: list[FolderEntry],
    dest_root:         Path,
    program_filter:    str,
    dry_run:           bool,
    timestamp:         str,
) -> tuple[Path, Path]:
    """
    Writes two report files to dest_root/Reports/:
      • reconciliation_<TAG>_<ts>.xlsx  — 4-sheet Excel workbook
      • missing_records_<TAG>_<ts>.txt  — plain-text list

    Returns (xlsx_path, txt_path).
    """
    report_dir = dest_root / "Reports"
    report_dir.mkdir(parents=True, exist_ok=True)

    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
    tag = "DRYRUN" if dry_run else "EXECUTED"

    matched = [r for r in results if r.folder is not None]
    missing = [r for r in results if r.folder is None]

    # ── Excel workbook ────────────────────────────────────────────────────────
    xlsx_path = report_dir / f"reconciliation_{tag}_{ts}.xlsx"
    wb = Workbook(); wb.remove(wb.active)

    # Summary
    ws = wb.create_sheet("📊 Summary"); ws.sheet_view.showGridLines = False
    stats = [
        ("Programme Filter",  program_filter.upper()),
        ("Run Mode",          "DRY RUN — no files copied" if dry_run else "EXECUTED — files copied"),
        ("Generated",         timestamp),
        ("", ""),
        ("Total Excel Records",  len(results)),
        ("Successful Matches",   len(matched)),
        ("  — Exact",            sum(1 for r in matched if r.match_type == "exact")),
        ("  — Token",            sum(1 for r in matched if r.match_type == "token")),
        ("  — Fuzzy",            sum(1 for r in matched if r.match_type == "fuzzy")),
        ("Missing Records",      len(missing)),
        ("Unmatched Folders",    len(unmatched_folders)),
        ("Match Rate",           f"{len(matched) / max(len(results), 1):.1%}"),
    ]
    lf = _fill("E3F2FD")
    for r_idx, (lbl, val) in enumerate(stats, 2):
        lc = ws.cell(row=r_idx, column=2, value=lbl)
        vc = ws.cell(row=r_idx, column=3, value=val)
        lc.font = _font(bold=True); vc.font = _font()
        lc.fill = lf; lc.border = _border(); vc.border = _border()
        ws.row_dimensions[r_idx].height = 20
    ws.column_dimensions["B"].width = 28; ws.column_dimensions["C"].width = 42

    # Matched
    ws = wb.create_sheet("✅ Matched"); ws.sheet_view.showGridLines = False
    _hdr(ws, ["#", "Student Name", "Programme", "Match Type", "Score",
               "Source Folder", "Destination"])
    for i, r in enumerate(matched, 1):
        fill = _fill("E8F5E9") if i % 2 == 0 else _fill("F9FAFB")
        for c, v in enumerate(
            [i, r.student.raw_name, r.student.raw_program,
             r.match_type.upper(), f"{r.match_score:.0%}",
             r.folder.path.name, str(r.dest_path) if r.dest_path else "—"], 1
        ):
            cell = ws.cell(row=i + 1, column=c, value=v)
            cell.fill = fill; cell.border = _border()
            cell.alignment = Alignment(vertical="center")
    _col_widths(ws, [5, 32, 46, 12, 8, 36, 54])
    ws.freeze_panes = "A2"

    # Missing
    ws = wb.create_sheet("❌ Missing Records"); ws.sheet_view.showGridLines = False
    _hdr(ws, ["#", "Student Name", "Surname", "First Name", "Programme", "Type", "Excel Row"])
    for i, r in enumerate(missing, 1):
        s = r.student
        for c, v in enumerate(
            [i, s.raw_name, s.surname_raw, s.firstname_raw,
             s.raw_program, s.program_type.upper(), s.row_num], 1
        ):
            cell = ws.cell(row=i + 1, column=c, value=v)
            cell.fill = _fill("FFEBEE"); cell.border = _border()
    _col_widths(ws, [5, 30, 22, 22, 46, 10, 10])
    ws.freeze_panes = "A2"

    # Unmatched folders
    ws = wb.create_sheet("⚠️ Unmatched Folders"); ws.sheet_view.showGridLines = False
    _hdr(ws, ["#", "Folder Name", "Normalised Name", "Action"])
    for i, f in enumerate(unmatched_folders, 1):
        for c, v in enumerate(
            [i, f.path.name, f.norm_name, "Not in Excel roster — verify manually"], 1
        ):
            cell = ws.cell(row=i + 1, column=c, value=v)
            cell.fill = _fill("FFF8E1"); cell.border = _border()
    _col_widths(ws, [5, 40, 40, 38])
    ws.freeze_panes = "A2"

    wb.save(xlsx_path)

    # ── Plain-text missing list ───────────────────────────────────────────────
    txt_path = report_dir / f"missing_records_{tag}_{ts}.txt"
    lines = [
        "MISSING RECORDS REPORT",
        f"Generated : {timestamp}",
        f"Filter    : {program_filter.upper()}",
        f"Mode      : {'DRY RUN' if dry_run else 'EXECUTED'}",
        f"Total     : {len(missing)} missing out of {len(results)} records",
        "=" * 62, "",
    ]
    for i, r in enumerate(missing, 1):
        lines.append(f"{i:>3}. {r.student.raw_name:<42}  {r.student.raw_program}")
    if not missing:
        lines.append("  *** All records matched — no missing folders! ***")
    txt_path.write_text("\n".join(lines), encoding="utf-8")

    return xlsx_path, txt_path
